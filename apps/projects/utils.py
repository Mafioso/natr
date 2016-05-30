#!/usr/bin/env python
# -*- coding: utf-8 -*-
from natr.settings import EXCEL_REPORTS_DIR
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, Color, colors, PatternFill
from openpyxl.cell import get_column_letter
from django.core.exceptions import ObjectDoesNotExist
from django.db.models.query import QuerySet
from projects import models as prj_models
from documents import models as doc_models
import decimal
import os
import dateutil.parser
import zipfile

def get_money_amount(field):
    if not field:
        return 0

    return field.amount

def get_value(value):
    if not value:
        return u""

    return value

class ExcelReport:
    assignments = []
    assig_intervals = [
        u'Никогда',
        u'Еженедельно',
        u'Ежемесячно',
        u'Ежеквартально',
        u'Каждые полгода',
        u'Ежегодно'
    ]
    assig_statuses =  [
        u'Не назначено',
        u'Назначено',
        u'Отправлено на доработку',
        u'Исполнение продолжается',
        u'На утверждении',
        u'Утверждено',
        u'Исполнено'
    ]
    alignment_left=Alignment(horizontal='left',
                    vertical='top',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
    alignment_right=Alignment(horizontal='right',
                    vertical='top',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
    alignment_center=Alignment(horizontal='center',
                    vertical='top',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)

    alignment_header=Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
    border = Border(left=Side(border_style='thin', color=Color(rgb='FF000000')),
                 right=Side(border_style='thin', color=Color(rgb='FF000000')),
                 top=Side(border_style='thin', color=Color(rgb='FF000000')),
                 bottom=Side(border_style='thin', color=Color(rgb='FF000000')))
    yellowFill = PatternFill(start_color='FFFFC001',
                   end_color='FFFFC001',
                   fill_type='solid')

    greyFill = PatternFill(start_color='B3B3B3',
                   end_color='B3B3B3',
                   fill_type='solid')

    def __init__(self, report=None, projects=None, registry_data = None):
        self.report = report
        self.projects = projects
        self.registry_data = registry_data

    def build_header_cell(self, ws, column, cell_number, string):
        ws = self.insert_into_cell(ws, column, cell_number, string)
        ws.alignment = self.alignment_center
        ws.font = Font(bold=True)
        return ws

    def set_border(self, ws, cell_range):
        thin_border = Border(
                                left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin')
                            )

        ws.cell(row=3, column=2).border = thin_border
        for row in range(cell_range['row_start'], cell_range['row_finish']):
            for col in range(cell_range['col_start'], cell_range['col_finish']):
                ws.cell(row=row, column=col).border = thin_border

    def insert_into_cell(self, ws, column, cell_number, string):
        if not string:
            string = ""
        if type(cell_number) == int:
            cell_number = str(cell_number)
        if type(string) == int:
            string = str(string)
        ws[column+cell_number] = string
        if ws.column_dimensions[column].width < len(string):
            ws.column_dimensions[column].width = len(string) + 3
        # ws[column+cell_number].border = self.border
        return ws


    def generate_excel_report(self):
        report = self.report
        project = report.project
        milestone = report.milestone
        wb = Workbook()
        ws1 = wb.active
        ws1.title = u'Общая информация'

        ws1 = self.insert_into_cell(ws1, 'A', '1', u'Дата отчета')
        ws1 = self.insert_into_cell(ws1, 'A', '2', u'Наименование грантополучателя')
        ws1 = self.insert_into_cell(ws1, 'A', '3', u'Номер и дата договора')
        ws1 = self.insert_into_cell(ws1, 'A', '4', u'Цель гранта')
        ws1 = self.insert_into_cell(ws1, 'A', '5', u'Сумма гранта')
        ws1 = self.insert_into_cell(ws1, 'A', '6', u'Период отчетности')
        ws1 = self.insert_into_cell(ws1, 'A', '7', u'Достигнутые результаты грантового проекта')
        ws1 = self.insert_into_cell(ws1, 'A', '8', u'Наименование охранного документа (в случае наличия)')
        ws1 = self.insert_into_cell(ws1, 'A', '9', u'Номер и дата выдачи охранного документа (в случае наличия)')

        ws1 = self.insert_into_cell(ws1, 'B', '1', report.date)
        try:
            ws1 = self.insert_into_cell(ws1, 'B', '2', project.organization_details.name)
        except:
            ws1 = self.insert_into_cell(ws1, 'B', '2', "")
        contract_number_date = "{0}, {1}".format(project.aggreement.document.number, project.aggreement.document.date_sign) if project is not None else ""
        ws1 = self.insert_into_cell(ws1, 'B', '3', contract_number_date)
        ws1 = self.insert_into_cell(ws1, 'B', '4', project.grant_goal if project is not None else "")
        try:
            ws1 = self.insert_into_cell(ws1, 'B', '5', "{0} {1}".format(milestone.fundings.amount, milestone.fundings.currency))
        except:
            ws1 = self.insert_into_cell(ws1, 'B', '5', "")
        ws1 = self.insert_into_cell(ws1, 'B', '6', report.period if project is not None else "")
        ws1 = self.insert_into_cell(ws1, 'B', '7', report.results if project is not None else "")



        ws2 = wb.create_sheet()
        ws2.title = u'Бюждетные средства'

        ws2 = self.insert_into_cell(ws2, 'A', '1', u'Номер п/п')
        ws2 = self.insert_into_cell(ws2, 'B', '1', u'Статьи затрат по договору')
        ws2 = self.insert_into_cell(ws2, 'C', '1', u'Документы по отчету грантополучателя за Этап № {0}'.format(milestone.number))
        ws2 = self.insert_into_cell(ws2, 'D', '1', "")
        ws2 = self.insert_into_cell(ws2, 'E', '1', "")
        ws2 = self.insert_into_cell(ws2, 'F', '1', "")
        ws2 = self.insert_into_cell(ws2, 'G', '1', "")
        ws2 = self.insert_into_cell(ws2, 'H', '1', "")
        ws2.merge_cells('C1:H1')
        ws2['C1'].alignment = self.alignment_center
        ws2.merge_cells('A1:A2')
        ws2['A1'].alignment = self.alignment_center
        ws2.merge_cells('B1:B2')
        ws2['B1'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'C', '2', u'Наименование предприятия')
        ws2['C2'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'D', '2', u'Вид документа')
        ws2['D2'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'E', '2', u'№')
        ws2['E2'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'F', '2', u'Дата')
        ws2['F2'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'G', '2', u'Приложение')
        ws2['G2'].alignment = self.alignment_center
        ws2 = self.insert_into_cell(ws2, 'H', '2', u'Сумма, тенге')
        ws2['H2'].alignment = self.alignment_center

        use_of_budget_doc = report.use_of_budget_doc
        init_row = 3
        '''
            item.cost_name
            item.id
        '''
        for item in use_of_budget_doc.items.all():
            end_row = init_row
            ws2 = self.insert_into_cell(ws2, 'A', init_row, item.id)
            ws2 = self.insert_into_cell(ws2, 'B', init_row, item.cost_name)
            for cost in item.costs.all():
                cost_init_row = end_row
                ws2 = self.insert_into_cell(ws2, 'H', cost_init_row, str(cost.costs))
                ws2 = self.insert_into_cell(ws2, 'C', cost_init_row, cost.budget_item.project.organization_details.name)
                for gp_doc in cost.gp_docs.all():
                    ws2 = self.insert_into_cell(ws2, 'D', end_row, gp_doc.document.get_status_cap())
                    ws2 = self.insert_into_cell(ws2, 'F', end_row, gp_doc.document.date_created.date().__str__())
                    ws2 = self.insert_into_cell(ws2, 'E', end_row, gp_doc.id)
                    end_row += 1
                # ws2.merge_cells("H{}:H{}".format(str(cost_init_row), str(end_row)))
            # ws2.merge_cells("A{}:A{}".format(str(init_row), str(end_row)))
            # ws2.merge_cells("B{}:B{}".format(str(init_row), str(end_row)))
            init_row = end_row

        ws3 = wb.create_sheet()
        ws3.title = u'Затраты'

        ws3 = self.insert_into_cell(ws3, 'A', '1', u"№ п/п")
        ws3 = self.insert_into_cell(ws3, 'B', '1', u"Наименование статей затрат по смете")
        ws3 = self.insert_into_cell(ws3, 'C', '1', u"Сумма бюджетных средств по смете")
        ws3 = self.insert_into_cell(ws3, 'D', '1', u"Израсходованная сумма")
        ws3 = self.insert_into_cell(ws3, 'E', '1', u"Остаток средств")
        ws3 = self.insert_into_cell(ws3, 'F', '1', u"Наименование подтверждающих документов")
        ws3 = self.insert_into_cell(ws3, 'G', '1', u"Примечание")

        row = 2
        for item in use_of_budget_doc.items.all():
            ws3 = self.insert_into_cell(ws3, 'A', row, item.id)
            ws3 = self.insert_into_cell(ws3, 'B', row, item.cost_type.name)
            milestone_costs = item.milestone.factmilestonecostrow_set.all()
            budget_sum = 0
            for milestone_cost in milestone_costs:
                budget_sum += milestone_cost.costs.amount
            ws3 = self.insert_into_cell(ws3, 'C', row, str(budget_sum))
            ws3 = self.insert_into_cell(ws3, 'D', row, str(item.total_expense.amount))
            ws3 = self.insert_into_cell(ws3, 'E', row, str(budget_sum - item.total_expense.amount))
            try:
                ws3 = self.insert_into_cell(ws3, 'F', row, item.documents[0].document.get_status_cap())
            except:
                ws3 = self.insert_into_cell(ws3, 'F', row, str(0))
            ws3 = self.insert_into_cell(ws3, 'G', row, item.notes)
            row += 1
        file_dir = EXCEL_REPORTS_DIR
        if not os.path.exists(EXCEL_REPORTS_DIR):
            os.makedirs(EXCEL_REPORTS_DIR)
        filename = EXCEL_REPORTS_DIR + '/' + u'Отчет ' + str(project.aggreement.document.number) + '.xlsx'
        wb.save(filename)
        return filename

    #projects registry
    def generate_experts_report(self):
        projects = self.projects
        max_milestone_num = 0
        for project in projects:
            if len(project.milestone_set.all())>max_milestone_num:
                max_milestone_num = len(project.milestone_set.all())

        wb = Workbook()
        ws = wb.active

        if len(projects) == 0:
            ws = self.insert_into_cell(ws, 'A', '1', u'Сводный отчет по грантам')
            ws = self.insert_into_cell(ws, 'A', '2', u'Проектов не найдено')
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

            file_dir = EXCEL_REPORTS_DIR
            if not os.path.exists(EXCEL_REPORTS_DIR):
                os.makedirs(EXCEL_REPORTS_DIR)

            filename = EXCEL_REPORTS_DIR+'/'+u'Отчет по грантам.xlsx'
            wb.save(filename)
            return filename

        all_cols_number = 0
        temp_col_num = 0
        first = True
        p_len = 0
        row = 1

        if type(projects) == list:
            p_len = len(projects)
        elif type(projects) == QuerySet:
            p_len = projects.count()

        for project, cnt in zip( self.projects, range(p_len) ):
            col_num = 1
            row = cnt+4
            if first:
                ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'№ п/п' )

            ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(cnt+1))

            col_num += 1

            if 'grantee_name' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Наименование ГП' )
                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.organization_details.name)
                except ObjectDoesNotExist:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            if 'project_name' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Наименование проекта')
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.name) 
                col_num += 1

            
            if 'project_description' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Краткое описание проекта')

                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.description)
                except: 
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                
                col_num += 1

            if 'project_innovation' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Инновационность проекта')

                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.innovation)
                except: 
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                
                col_num += 1

            if 'problem_questions' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Проблемные вопросы')

                problem_questions = ""

                for question in project.problem_questions.all():
                    problem_questions += question.name + "\n"

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, problem_questions)
                
                col_num += 1

            if 'grant_type' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Вид гранта' )
                if project.funding_type:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.funding_type.get_name_display())
                col_num += 1

            if 'region' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Регион' )
                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.organization_details.get_address_region_display())
                except:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            if 'contact_details' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Контактные данные' )
                try:
                    contact_details = project.organization_details.contact_details
                    project_contact_details = project.organization_details.address_1 + "\n" + \
                                              contact_details.phone_number + "\n" + \
                                              contact_details.email
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project_contact_details)
                except:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            if 'bin' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'БИН или ИИН' )
                try:
                    bin_details = project.organization_details.bin + "\n" + \
                                              project.organization_details.bik + "\n" + \
                                              project.organization_details.iik

                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, bin_details)
                except:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            if 'aggreement' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'№/дата договора' )

                if project.aggreement: 
                    agreement_number = project.aggreement.document.number if project.aggreement.document.number else ""
                else:
                    agreement_number = ""

                if project.aggreement and project.aggreement.document.date_sign:
                    agreement_date = u' от ' + project.aggreement.document.date_sign.strftime("%d.%m.%y")
                else:
                    agreement_date = ""
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, agreement_number + agreement_date)
                
                col_num += 1
                
            if 'other_agreements' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'№ и дата доп. согл.' )

                try:
                    other_agreements = project.other_agreements
                    items = []
                    if hasattr(other_agreements, "items"):
                        for item in other_agreements.items.all():
                            date_sign = ""
                            if item.date_sign:
                                date_sign = u' от ' + item.date_sign.strftime("%d.%m.%y")

                            items.append(item.number + date_sign)

                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "\n".join(items))
                except ObjectDoesNotExist:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')

                col_num += 1

            if 'grant_period' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Период предоставления гранта' )
                
                agreement_date = ""
                if project.aggreement and project.aggreement.document.date_sign:
                    agreement_date = project.aggreement.document.date_sign.year

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, agreement_date)
                col_num += 1

            if 'total_month' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Срок реализации проекта, (мес)' )
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.total_month)
                col_num += 1

            if 'fundings' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Сумма по Договору, (тенге)')
                sum_fundings = "0"
                if project.aggreement and project.aggreement.funding:
                    sum_fundings = project.aggreement.funding.amount

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(sum_fundings))
                col_num += 1

            if 'natr_fundings' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Сумма гранта')
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(project.fundings.amount) if project.fundings else "0" )
                col_num += 1

            if 'own_fundings' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Cобственные средства')

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(project.own_fundings.amount) if project.own_fundings else "0" )
                col_num += 1

            if 'expert' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Исполнитель' )
                assigned_experts = []

                for expert in project.assigned_experts.all():
                    assigned_experts.append(expert.get_full_name())

                ws = self.build_header_cell( ws, get_column_letter(col_num), row, ",".join(assigned_experts) )
                col_num += 1

            if 'status' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Статус ' )     
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.get_status_cap())               
                col_num += 1

            if 'number_of_milesones' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Количество этапов' )
                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.number_of_milestones)
                except:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            money_sum = 0
            temp_col_num = col_num
            for milestone in project.milestone_set.all().order_by('number'):
                if milestone.fundings:
                    money_sum += milestone.fundings.amount

                if 'transhes' in self.registry_data['keys']: 
                    try:
                        ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(milestone.fundings.amount))
                    except:
                        ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(0))
                    col_num += 1
            
            if 'transhes' in self.registry_data['keys']: 
                if first:
                    for i in range(temp_col_num, temp_col_num+max_milestone_num):
                        ws = self.build_header_cell(ws, get_column_letter(i), row-1, u'{} транш'.format(str(i-temp_col_num+1)) )
                col_num = temp_col_num + max_milestone_num    
                
            if 'balance' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Остаток денежных средств,  (тенге)' )
                if not project.fundings:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(0))
                else:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(project.fundings.amount - money_sum))
                col_num += 1

            if 'total_fundings' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Итого перечислено')
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(money_sum))
                col_num += 1

            if 'gp_type' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Вид ГП' )
                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.organization_details.get_org_type_display())
                except ObjectDoesNotExist:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1

            if 'risks' in self.registry_data['keys']:
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'Риски' )
                try:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.risk_title)
                except:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1
            

            first = False
            
        all_cols_number = col_num
        ws = self.insert_into_cell(ws, 'A', '1', u'Отчет по грантам')
        ws.merge_cells('A1:%s1'%get_column_letter(all_cols_number-1 if all_cols_number > 1 else 1))
        ws['A1'].alignment = self.alignment_center
        ws['A1'].font = Font(bold=True)

        row += 3

        ws = self.insert_into_cell(ws, 'A', row, u'ВСЕГО по Договорам 2015 года:')
        ws.merge_cells("A%s:%s%s"%(row, get_column_letter(temp_col_num-1 if temp_col_num > 2 else 1), row))
        ws["A{}".format(row)].alignment = self.alignment_right

        last_cols = 0
        if 'gp_type' in self.registry_data['keys']:
            last_cols += 1
        if 'risks' in self.registry_data['keys']:
            last_cols += 1

        if temp_col_num > 2 and all_cols_number > 2:
            for i in range(temp_col_num, all_cols_number - last_cols):
                sum = 0
                for j in range(2, row-1):
                    val = ws['{}{}'.format(get_column_letter(i), j)].value
                    try:                 
                        sum += decimal.Decimal(val)
                    except:
                        pass
                ws = self.insert_into_cell(ws, get_column_letter(i), row, str(sum))
        
        greyFill = PatternFill(start_color='B3B3B3',
                   end_color='B3B3B3',
                   fill_type='solid')

        for i in range(all_cols_number-1):
            ws.column_dimensions[get_column_letter(i+1)].width = 20.0
            ws[get_column_letter(i+1)+"3"].font = Font(bold=True)
            ws[get_column_letter(i+1)+"3"].fill = greyFill

        self.set_border(ws, {'col_start': 1,
                             'col_finish': all_cols_number, 
                             'row_start': 3,
                             'row_finish': 7+p_len})

        file_dir = EXCEL_REPORTS_DIR
        if not os.path.exists(EXCEL_REPORTS_DIR):
            os.makedirs(EXCEL_REPORTS_DIR)

        filename = EXCEL_REPORTS_DIR+'/'+u'Отчет по грантам 2015 года.xlsx'

        if 'date_from' in self.registry_data and "date_to" in self.registry_data:
            dates = self.registry_data['date_from'].strftime("%d.%m.%y") + \
                    "-" + self.registry_data['date_to'].strftime("%d.%m.%y")
            
            ws = self.insert_into_cell(ws, 'A', '1', u'Сводный отчет по грантам за период' + dates)
            ws.merge_cells('A1:%s1'%get_column_letter(all_cols_number-1 if all_cols_number > 1 else 1))
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

            ws = self.insert_into_cell(ws, 'A', row, u'ВСЕГО по Договорам ' + dates)
            ws.merge_cells("A%s:%s%s"%(row, get_column_letter(temp_col_num-2 if temp_col_num > 2 else 1), row))
            ws["A{}".format(row)].alignment = self.alignment_right

            filename = EXCEL_REPORTS_DIR + '/' + u'Реестр проектов ' + dates + '.xlsx'


        wb.save(filename)
        return filename

    #project efficiency report
    def generate_efficiency_report(self):
        wb = Workbook()
        ws = wb.active
        projects = self.projects

        file_dir = EXCEL_REPORTS_DIR
        if not os.path.exists(EXCEL_REPORTS_DIR):
            os.makedirs(EXCEL_REPORTS_DIR)

        if len(projects) == 0:
            ws = self.insert_into_cell(ws, 'A', '1', u'Оценка эффективности реализации проектов')
            ws = self.insert_into_cell(ws, 'A', '2', u'Проектов не найдено')
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

            filename = EXCEL_REPORTS_DIR+'/'+u'Оценка эффективности реализации проектов.xlsx'
            wb.save(filename)
            return filename

        date_from = dateutil.parser.parse(self.registry_data['date_from'])
        date_to = dateutil.parser.parse(self.registry_data['date_to'])

        all_cols_number = 0
        temp_col_num = 0
        col_num = 1
        first = True
        p_len = 0
        row = 7
        at_least_one = False

        if type(projects) == list:
            p_len = len(projects)
        elif type(projects) == QuerySet:
            p_len = projects.count()


        for project, cnt in zip( self.projects, range(p_len) ):
            efficiency_objs = project.get_efficiency_objs_in_period(date_from, date_to)
            first_eff_obj = True
            start_row = row
            for efficiency_obj in efficiency_objs:
                col_num = 1
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'№ п/п' )

                if first_eff_obj:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(cnt+1))

                col_num += 1

                #region aggreement date and number
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'№/дата договора' )

                if first_eff_obj:
                    if project.aggreement: 
                        agreement_number = project.aggreement.document.number if project.aggreement.document.number else ""
                    else:
                        agreement_number = ""

                    if project.aggreement and project.aggreement.document.date_sign:
                        agreement_date = u' от ' + project.aggreement.document.date_sign.strftime("%d.%m.%y")
                    else:
                        agreement_date = ""
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, agreement_number + agreement_date)
                
                col_num += 1
                #endregion

                #region grantee name
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Наименование грантополучателя' )
                
                if first_eff_obj:
                    try:
                        ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.organization_details.name)
                    except ObjectDoesNotExist:
                        ws = self.insert_into_cell(ws, get_column_letter(col_num), row, '')
                col_num += 1
                #endregion

                #region project status
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Статус проекта ' )     
                if first_eff_obj:
                    ws = self.insert_into_cell(ws, get_column_letter(col_num), row, project.get_status_cap())               
                col_num += 1
                #endregion

                #region directors conclusion
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Решение Правления Агентства ' )  
                col_num += 1
                #endregion

                #region directors conclusion
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Дата Отчета' ) 

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, efficiency_obj.report_date.strftime("%d.%m.%y") if efficiency_obj.report_date else "") 
                col_num += 1
                #endregion



                #region created job places
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Количество создан-х раб.мест ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.workplaces_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.workplaces_fact))) 
                col_num += 1
                #endregion

                #region product types
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Количество видов производимой продукции, шт ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.types_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.types_fact))) 
                col_num += 1
                #endregion

                #region product names
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Наименование вида выпускаемой продукции ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                #endregion

                #region product volume
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Объем выпущенной продукции, тенге ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'в натуральном выражении (шт./тонн/литр и т.п.) ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-2, u'в денежном выражении, тенге ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-1, u'факт ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+3), row-1, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.prod_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.prod_fact))) 
                col_num += 1
                #endregion

                #region realized volume
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Объем реализованной продукции, тенге ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'всего ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-2, u'на внутренний рынок ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+4), row-2, u'на внешний рынок (экспорт) ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-1, u'факт ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+3), row-1, u'факт ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+4), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+5), row-1, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.total_rlzn_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.total_rlzn_fact))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.rlzn_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.rlzn_fact))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.rlzn_exp_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.rlzn_exp_fact))) 
                col_num += 1
                #endregion

                #region taxes volume
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Сумма уплаченных налоговых средств, тенге ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'всего ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-2, u'в республиканский бюджет ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+4), row-2, u'в местный бюджет ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-1, u'факт ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+2), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+3), row-1, u'факт ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+4), row-1, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+5), row-1, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.total_tax_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.total_tax_fact))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.tax_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.tax_fact))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.tax_local_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_money_amount(efficiency_obj.tax_local_fact))) 
                col_num += 1
                #endregion

                #region crated innovative products
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Количество внедренных инновационных продуктов, ед. ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.innovs_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.innovs_fact))) 
                col_num += 1
                #endregion

                #region kazakh market products share
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Доля казахстанского содержания в продукци, работах и услугах, % ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.kaz_part_plan))) 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, str(get_value(efficiency_obj.kaz_part_fact))) 
                col_num += 1
                #endregion

                #region safety documents
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Наименование охранных документов (при наличии) ' )     

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "")               
                col_num += 1
                #endregion

                #region project power
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Проектная мощность, в натуральном выражении (шт./тонн/литр и т.п.) ' ) 
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-2, u'план ' )
                    ws = self.build_header_cell(ws, get_column_letter(col_num+1), row-2, u'факт ' )

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "") 
                col_num += 1
                #endregion

                #region performance index
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Индекс производительности труда по итогам получения инновационного гранта, в процентах** ' )     

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "")               
                col_num += 1
                #endregion

                #region results
                if first:
                    ws = self.build_header_cell(ws, get_column_letter(col_num), row-3, u'Результат по итогам реализации проекта (технология, новый вид продукции и т.п.) ' )     

                ws = self.insert_into_cell(ws, get_column_letter(col_num), row, "")               
                col_num += 1
                #endregion

                first = False
                first_eff_obj = False
                at_least_one = True
                row += 1

            ws.merge_cells("A%s:A%s"%(start_row, row-1))
            ws.merge_cells("B%s:B%s"%(start_row, row-1))
            ws.merge_cells("C%s:C%s"%(start_row, row-1))
            ws.merge_cells("D%s:D%s"%(start_row, row-1))
            ws.merge_cells("E%s:E%s"%(start_row, row-1))

        if at_least_one:
            greyFill = PatternFill(start_color='B3B3B3',
                       end_color='B3B3B3',
                       fill_type='solid')

            all_cols_number = col_num
            for i in range(all_cols_number-1):
                ws.column_dimensions[get_column_letter(i+1)].width = 20.0
                ws[get_column_letter(i+1)+"4"].font = Font(bold=True)
                ws[get_column_letter(i+1)+"4"].fill = greyFill
                ws[get_column_letter(i+1)+"5"].font = Font(bold=True)
                ws[get_column_letter(i+1)+"5"].fill = greyFill
                ws[get_column_letter(i+1)+"6"].font = Font(bold=True)
                ws[get_column_letter(i+1)+"6"].fill = greyFill
                ws.merge_cells('A4:A6')
                ws.merge_cells('B4:B6')
                ws.merge_cells('C4:C6')
                ws.merge_cells('D4:D6')
                ws.merge_cells('E4:E6')
                ws.merge_cells('F4:F6')
                ws.merge_cells('G4:H4')
                ws.merge_cells('G5:G6')
                ws.merge_cells('H5:H6')
                ws.merge_cells('I4:J4')
                ws.merge_cells('I5:I6')
                ws.merge_cells('J5:J6')
                ws.merge_cells('K4:L4')
                ws.merge_cells('K5:K6')
                ws.merge_cells('L5:L6')
                ws.merge_cells('M4:P4')
                ws.merge_cells('M5:N5')
                ws.merge_cells('O5:P5')
                ws.merge_cells('Q4:V4')
                ws.merge_cells('Q5:R5')
                ws.merge_cells('S5:T5')
                ws.merge_cells('U5:V5')
                ws.merge_cells('W4:AB4')
                ws.merge_cells('W5:X5')
                ws.merge_cells('Y5:Z5')
                ws.merge_cells('AA5:AB5')
                ws.merge_cells('AC4:AD4')
                ws.merge_cells('AC5:AC6')
                ws.merge_cells('AD5:AD6')
                ws.merge_cells('AE4:AF4')
                ws.merge_cells('AE5:AE6')
                ws.merge_cells('AF5:AF6')
                ws.merge_cells('AG4:AG6')
                ws.merge_cells('AH4:AI4')
                ws.merge_cells('AH5:AH6')
                ws.merge_cells('AI5:AI6')
                ws.merge_cells('AJ4:AJ6')
                ws.merge_cells('AK4:AK6')

            self.set_border(ws, {'col_start': 1,
                                 'col_finish': all_cols_number, 
                                 'row_start': 4,
                                 'row_finish': row})
        else:
            ws = self.insert_into_cell(ws, 'A', '1', u'Оценка эффективности реализации проектов по итогам')
            ws = self.insert_into_cell(ws, 'A', '2', u'Перепроверьте данные в отчете, за данный период отчетов не найдено')
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

        filename = EXCEL_REPORTS_DIR+'/'+u'Отчет по эффективности проектов.xlsx'

        date_from = dateutil.parser.parse(self.registry_data['date_from']) if self.registry_data['date_from'] else None
        date_to = dateutil.parser.parse(self.registry_data['date_to']) if self.registry_data['date_to'] else None

        if date_from and date_to:
            dates = date_from.strftime("%d.%m.%y") + "-" + date_to.strftime("%d.%m.%y")
            
            ws = self.insert_into_cell(ws, 'A', '1', u'Оценка эффективности реализации проектов по итогам ' + dates)
            ws.merge_cells('A1:%s1'%get_column_letter(all_cols_number-1 if all_cols_number > 1 else 1))
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

            filename = EXCEL_REPORTS_DIR + '/' + u'Отчет по эффективности проектов ' + dates + '.xlsx'

        wb.save(filename)
        return filename

    #comments and archive report
    def generate_archive_report(self):
        wb = Workbook()
        ws = wb.active
        projects = self.projects

        if len(projects) == 0:
            ws = self.insert_into_cell(ws, 'A', '1', u'Архив сообщений и комментариев')
            ws = self.insert_into_cell(ws, 'A', '2', u'Проектов не найдено')
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = self.alignment_center
            ws['A1'].font = Font(bold=True)

            file_dir = EXCEL_REPORTS_DIR
            if not os.path.exists(EXCEL_REPORTS_DIR):
                os.makedirs(EXCEL_REPORTS_DIR)

            filename = EXCEL_REPORTS_DIR+'/'+u'Архив сообщений и комментариев.xlsx'
            wb.save(filename)
            return filename

        insert_header = True
        date_from = self.registry_data['date_from']
        date_to = self.registry_data['date_to']
        dates = date_from.strftime("%d.%m.%y") + "-" + date_to.strftime("%d.%m.%y")
        active_wb = True
        chat_ws = None
        pm_ws = None
        report_ws = None
        corollary_ws = None

        greyFill = PatternFill(start_color='B3B3B3',
                       end_color='B3B3B3',
                       fill_type='solid')


        def fill_chat_tab(project, row):
            if insert_header:
                chat_ws = wb.create_sheet() if  not active_wb else wb.active
                chat_ws.title = u"Сообщения из чата по проектам"
                chat_ws = self.insert_into_cell(chat_ws, 'A', '1', u'Сообщения из чата по проектам в период %s'%dates)
                chat_ws.merge_cells('A1:C1')
                chat_ws = self.insert_into_cell(chat_ws, 'A', row, u'Дата')
                chat_ws = self.insert_into_cell(chat_ws, 'B', row, u'ФИО Отправителя')
                chat_ws = self.insert_into_cell(chat_ws, 'C', row, u'Текст сообщения')
                chat_ws["A%i"%row].fill = greyFill
                chat_ws["B%i"%row].fill = greyFill
                chat_ws["C%i"%row].fill = greyFill
                row += 1

            chat_ws = self.insert_into_cell(chat_ws, 'A', row, u'Проект: %s'%project.name)
            chat_ws.merge_cells('A%i:C%i'%(row, row))
            row += 1

            messages = project.textline_set.filter(date_created__gte=date_from, date_created__lte=date_to)

            if messages.count() == 0:
                chat_ws = self.insert_into_cell(chat_ws, 'A', row, u'Сообщений не найдено')
                chat_ws.merge_cells('A%i:C%i'%(row, row))
                row += 1

            for chat_message in messages:
                chat_ws = self.insert_into_cell(chat_ws, 'A', row, chat_message.date_created.strftime("%d.%m.%y"))
                chat_ws = self.insert_into_cell(chat_ws, 'B', row, chat_message.from_account.get_full_name())
                chat_ws = self.insert_into_cell(chat_ws, 'C', row, chat_message.line)
                row += 1


            return row, chat_ws

        def fill_pm_comments_tab(project, row):
            if insert_header:
                pm_ws = wb.create_sheet() if  not active_wb else wb.active
                pm_ws.title = u"Комментарии к плану мониторинга"
                pm_ws = self.insert_into_cell(pm_ws, 'A', '1', u'Комментарии к плану мониторинга в период %s'%dates)
                pm_ws.merge_cells('A1:C1')
                pm_ws = self.insert_into_cell(pm_ws, 'A', row, u'Дата')
                pm_ws = self.insert_into_cell(pm_ws, 'B', row, u'ФИО Отправителя')
                pm_ws = self.insert_into_cell(pm_ws, 'C', row, u'Текст комментария')
                pm_ws["A%i"%row].fill = greyFill
                pm_ws["B%i"%row].fill = greyFill
                pm_ws["C%i"%row].fill = greyFill
                row += 1

            pm_ws = self.insert_into_cell(pm_ws, 'A', row, u'Проект: %s'%project.name)
            pm_ws.merge_cells('A%i:C%i'%(row, row))
            row += 1

            for monitoring in project.monitoring_set.all():
                comments = monitoring.comments.filter(date_created__gte=date_from, date_created__lte=date_to)

                if comments.count() == 0:
                    pm_ws = self.insert_into_cell(pm_ws, 'A', row, u'Комментариев не найдено')
                    pm_ws.merge_cells('A%i:C%i'%(row, row))
                    row += 1

                for comment in comments:
                    pm_ws = self.insert_into_cell(pm_ws, 'A', row, comment.date_created.strftime("%d.%m.%y")) 
                    pm_ws = self.insert_into_cell(pm_ws, 'B', row, comment.account.get_full_name())
                    pm_ws = self.insert_into_cell(pm_ws, 'C', row, comment.comment_text)
                    row += 1

            return row, pm_ws

        def fill_report_comments_tab(project, row):
            if insert_header:
                report_ws = wb.create_sheet() if  not active_wb else wb.active
                report_ws.title = u"Комментарии к отчету"
                report_ws = self.insert_into_cell(report_ws, 'A', '1', u'Комментарии к отчету в период %s'%dates)
                report_ws.merge_cells('A1:C1')
                report_ws = self.insert_into_cell(report_ws, 'A', row, u'Дата')
                report_ws = self.insert_into_cell(report_ws, 'B', row, u'ФИО Отправителя')
                report_ws = self.insert_into_cell(report_ws, 'C', row, u'Текст комментария')
                report_ws["A%i"%row].fill = greyFill
                report_ws["B%i"%row].fill = greyFill
                report_ws["C%i"%row].fill = greyFill
                row += 1

            report_ws = self.insert_into_cell(report_ws, 'A', row, u'Проект: %s'%project.name)
            report_ws.merge_cells('A%i:C%i'%(row, row))
            row += 1

            for report in project.report_set.all().order_by("milestone__number"):
                report_ws = self.insert_into_cell(report_ws, 'A', row, u'Комментарии к отчету этапа %i'%report.milestone.number)
                report_ws.merge_cells('A%i:C%i'%(row, row))
                row += 1

                comments = report.comments.filter(date_created__gte=date_from, date_created__lte=date_to)

                if comments.count() == 0:
                    report_ws = self.insert_into_cell(report_ws, 'A', row, u'Комментариев не найдено')
                    report_ws.merge_cells('A%i:C%i'%(row, row))
                    row += 1
                for comment in comments:
                    report_ws = self.insert_into_cell(report_ws, 'A', row, comment.date_created.strftime("%d.%m.%y")) 
                    report_ws = self.insert_into_cell(report_ws, 'B', row, comment.account.get_full_name())
                    report_ws = self.insert_into_cell(report_ws, 'C', row, comment.comment_text)
                    row += 1

            return row, report_ws 

        def fill_corollary_comments_tab(project, row):
            if insert_header:
                corollary_ws = wb.create_sheet() if  not active_wb else wb.active
                corollary_ws.title = u"Комментарии к заключению"
                corollary_ws = self.insert_into_cell(corollary_ws, 'A', '1', u'Комментарии к заключению в период %s'%dates)
                corollary_ws.merge_cells('A1:C1')
                corollary_ws = self.insert_into_cell(corollary_ws, 'A', row, u'Дата')
                corollary_ws = self.insert_into_cell(corollary_ws, 'B', row, u'ФИО Отправителя')
                corollary_ws = self.insert_into_cell(corollary_ws, 'C', row, u'Текст комментария')
                corollary_ws["A%i"%row].fill = greyFill
                corollary_ws["B%i"%row].fill = greyFill
                corollary_ws["C%i"%row].fill = greyFill
                row += 1

            corollary_ws = self.insert_into_cell(corollary_ws, 'A', row, u'Проект: %s'%project.name)
            corollary_ws.merge_cells('A%i:C%i'%(row, row))
            row += 1

            for corollary in project.corollary_set.all().order_by("milestone__number"):
                corollary_ws = self.insert_into_cell(corollary_ws, 'A', row, u'Комментарии к заключению этапа %i'%corollary.milestone.number)
                corollary_ws.merge_cells('A%i:C%i'%(row, row))
                row += 1

                comments = corollary.comments.filter(date_created__gte=date_from, date_created__lte=date_to)

                if comments.count() == 0:
                    corollary_ws = self.insert_into_cell(corollary_ws, 'A', row, u'Комментариев не найдено')
                    corollary_ws.merge_cells('A%i:C%i'%(row, row))
                    row += 1

                for comment in comments:
                    corollary_ws = self.insert_into_cell(corollary_ws, 'A', row, comment.date_created.strftime("%d.%m.%y")) 
                    corollary_ws = self.insert_into_cell(corollary_ws, 'B', row, comment.account.get_full_name())
                    corollary_ws = self.insert_into_cell(corollary_ws, 'C', row, comment.comment_text)
                    row += 1

            return row, corollary_ws

        row_chat = 3
        row_pm = 3
        row_report = 3
        row_corollary = 3

        for project in projects:
            if 'chat' in self.registry_data['keys']:
                row_chat, chat_ws = fill_chat_tab(project, row_chat)
                active_wb = False

            if 'monitoring_plan' in self.registry_data['keys']:
                row_pm, pm_ws = fill_pm_comments_tab(project, row_pm)
                active_wb = False

            if 'report' in self.registry_data['keys']:
                row_report, report_ws = fill_report_comments_tab(project, row_report)
                active_wb = False

            if 'corollary' in self.registry_data['keys']:
                row_corollary, corollary_ws = fill_corollary_comments_tab(project, row_corollary)
                active_wb = False

            insert_header = False

        if 'chat' in self.registry_data['keys']:
            self.set_border(chat_ws, {'col_start': 1,
                                 'col_finish': 4, 
                                 'row_start': 3,
                                 'row_finish': row_chat})

        if 'monitoring_plan' in self.registry_data['keys']:
            self.set_border(pm_ws, {'col_start': 1,
                                 'col_finish': 4, 
                                 'row_start': 3,
                                 'row_finish': row_pm})


        if 'report' in self.registry_data['keys']:
            self.set_border(report_ws, {'col_start': 1,
                                 'col_finish': 4, 
                                 'row_start': 3,
                                 'row_finish': row_report})


        if 'corollary' in self.registry_data['keys']:
            self.set_border(corollary_ws, {'col_start': 1,
                                 'col_finish': 4, 
                                 'row_start': 3,
                                 'row_finish': row_corollary})


        file_dir = EXCEL_REPORTS_DIR
        if not os.path.exists(EXCEL_REPORTS_DIR):
            os.makedirs(EXCEL_REPORTS_DIR)

        filename = EXCEL_REPORTS_DIR+'/'+u'Архив сообщений и комментариев.xlsx'

        wb.save(filename)
        return filename


def resetSignature(report):
    if(report.signature.all()):
        report.signature.first().delete()

    return report

def create_use_of_budget_files_zip(report):
    file_dir = EXCEL_REPORTS_DIR
    if not os.path.exists(EXCEL_REPORTS_DIR):
        os.makedirs(EXCEL_REPORTS_DIR)

    filename = EXCEL_REPORTS_DIR+'/'+u'Проект №%s файлы отчета использованных средств.zip'%report.project.aggreement.document.number

    with zipfile.ZipFile(filename, 'w') as myzip:
        for use_of_budget_item in report.use_of_budget_doc.items.all():
            if hasattr(use_of_budget_item, "costs"):
                if use_of_budget_item.costs.count() > 0:
                    for cost in use_of_budget_item.costs.all():
                        if hasattr(cost, "gp_docs"):
                            if cost.gp_docs.count() > 0:
                                for gp_doc in cost.gp_docs.all():
                                    if hasattr(gp_doc.document, "attachments"):
                                        for attachment in gp_doc.document.attachments.all():
                                            myzip.write(attachment.file_path, os.path.basename(attachment.file_path), compress_type = zipfile.ZIP_DEFLATED)

    return filename